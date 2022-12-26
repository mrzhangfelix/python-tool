# -*- coding: utf-8 -*-
import openpyxl

excelName = 'D:\\Users\\felix\\code\\mytool\\excelTool\\jar.xlsx'
excelName2 = 'D:\\Users\\felix\\code\\mytool\\excelTool\\jar2.xlsx'
strlist=['HikariCP-2.7.9.jar',
'animal-sniffer-annotations-1.14.jar',
'byte-buddy-1.7.11.jar',
'checker-compat-qual-2.0.0.jar',
'classmate-1.3.4.jar',
'commons-collections-3.2.1.j',
'commons-collections-3.2.2.jar',
'commons-collections4-4.1.jar',
'commons-io-2.5.jar',
'commons-lang-2.6.jar',
'commons-lang3-3.7.jar',
'commons-logging-1.1.1.jar',
'commons-logging-commons-logging-1.2.jar',
'ezmorph-1.0.6.jar',
'fastjson-1.2.40.jar',
'guava-25.0-jre.jar',
'httpclient-4.5.5.jar',
'httpcore-4.4.9.jar',
'json-lib-2.4.jar',
'log4j-1.2.17.jar',
'lombok-1.16.22.jar',
'micrometer-core-1.0.10.jar',
'mybatis-3.4.0.jar',
'mybatis-spring-1.3.0.jar',
'mybatis-spring-boot-autoconfigure-1.1.1.jar',
'mybatis-spring-boot-starter-1.1.1.jar',
'org.apache.commons-commons-logging-1.2.jar',
'swagger-annotations-1.5.20.jar',
'swagger-models-1.5.20.jar',
'tomcat-embed-core-8.5.39.jar',
'tomcat-embed-el-8.5.39.jar',
'tomcat-embed-websocket-8.5.39.jar',
'xmlbeans-2.6.0.jar']

# 将path的excel中单元格中的字段包含在strlist中的复制到path2中去
def dowork(path,strlist,path2):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Sheet1"]

    wb2 = openpyxl.Workbook()
    sheet2 = wb2.active
    sheet2.title = '测试'

    for row in sheet.rows:
        for cell in row:
            if cell.value in strlist:
                sheet2.cell(row=cell.row, column=cell.column, value=cell.value)             
    wb2.save(path2)

def main():
    dowork(excelName,strlist,excelName2)

if __name__ == "__main__":
    main()
